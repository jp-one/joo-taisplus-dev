from odoo import models, fields
import openpyxl
import xlrd
import base64
from datetime import datetime
from io import BytesIO


class PriceListImport(models.TransientModel):
    _name = "taisplus.pricelist.import"
    _description = "Import Price List"

    file = fields.Binary(string="ファイル", required=True)
    filename = fields.Char(string="Filename")

    def import_pricelist(self):
        pricelist_model = self.env["taisplus.pricelist"].sudo()
        item_model = self.env["taisplus.pricelist.item"].sudo()

        # Validate filename format
        filename = self.filename or ""
        if not filename.lower().startswith("pricelist"):
            raise ValueError("Filename must start with 'pricelist'.")
        if not (
            filename.lower().endswith(".xlsx") or filename.lower().endswith(".xls")
        ):
            raise ValueError("Filename must have an extension of .xlsx or .xls.")

        # Extract date and expected headers
        try:
            if "_" in filename:
                # pricelist_YYYY-MM-DD_○○.xlsx or .xls
                date_str = filename.split("_")[1]
                header_date = datetime.strptime(date_str, "%Y-%m-%d").date()
                expected_headers = [
                    "商品コード",
                    "法人名",
                    "商品名",
                    "型番",
                    "全国平均貸与価格（円）",
                    "貸与価格の上限（円）",
                ]
            else:
                # pricelistYYYYMM.xlsx or .xls
                date_str = filename[9:15]
                header_date = datetime.strptime(date_str, "%Y%m").replace(day=1).date()
                expected_headers = [
                    "コード",
                    "法人名",
                    "商品名",
                    "型番",
                    "全国平均貸与価格（円）",
                    "貸与価格の上限（円）",
                ]
        except (IndexError, ValueError):
            raise ValueError(
                "Filename must include a valid date in 'YYYY-MM-DD' or 'YYYYMM' format "
                "(e.g., 'pricelist_2025-04-01_<remarks>.xlsx' or 'pricelist202504.xlsx')."
            )

        file_content = base64.b64decode(self.file)
        num_columns = len(expected_headers)
        temp_data = []
        titles = []
        sheet_names = ""

        if filename.lower().endswith(".xlsx"):
            workbook = openpyxl.load_workbook(BytesIO(file_content))
            sheet_names = ", ".join(sheet.title for sheet in workbook.worksheets)
            for sheet in workbook.worksheets:
                header_row = None
                for row_idx in range(1, 7):
                    actual_headers = [
                        sheet.cell(row=row_idx, column=col).value
                        for col in range(1, num_columns + 1)
                    ]
                    if actual_headers == expected_headers:
                        header_row = row_idx
                        break
                    titles.append(actual_headers)
                if header_row is None:
                    raise ValueError(
                        f"Valid headers not found in sheet '{filename!r}!{sheet.title}'. Expected: {expected_headers}"
                    )
                for row in sheet.iter_rows(
                    min_row=header_row + 1,
                    max_row=sheet.max_row,
                    min_col=1,
                    max_col=num_columns,
                ):
                    if row[0].value is None:
                        break
                    temp_data.append(
                        {
                            "tais_code": row[0].value,
                            "manufacturer": row[1].value,
                            "product_name": row[2].value,
                            "model_number": row[3].value,
                            "average_price": row[4].value,
                            "price_cap": row[5].value,
                        }
                    )
        else:  # .xls
            workbook = xlrd.open_workbook(file_contents=file_content)
            sheet_names = ", ".join(sheet.name for sheet in workbook.sheets())
            for sheet in workbook.sheets():
                header_row = None
                for row_idx in range(6):
                    actual_headers = [
                        sheet.cell_value(row_idx, col) for col in range(num_columns)
                    ]
                    if actual_headers == expected_headers:
                        header_row = row_idx
                        break
                    titles.append(actual_headers)
                if header_row is None:
                    raise ValueError(
                        f"Valid headers not found in sheet '{filename!r}!{sheet.name}'. Expected: {expected_headers}"
                    )
                for row_idx in range(header_row + 1, sheet.nrows):
                    if not sheet.cell_value(row_idx, 0):
                        break
                    temp_data.append(
                        {
                            "tais_code": sheet.cell_value(row_idx, 0),
                            "manufacturer": sheet.cell_value(row_idx, 1),
                            "product_name": sheet.cell_value(row_idx, 2),
                            "model_number": sheet.cell_value(row_idx, 3),
                            "average_price": sheet.cell_value(row_idx, 4),
                            "price_cap": sheet.cell_value(row_idx, 5),
                        }
                    )

        notes = "\n".join(
            " ".join(str(cell) for cell in row if cell) for row in titles if any(row)
        )

        header_name = f"上限一覧 {header_date.strftime('%Y-%m-%d')}"
        header_rec = pricelist_model.search(
            [("tais_code_date", "=", header_date)], limit=1
        )
        if not header_rec:
            header_rec = pricelist_model.create(
                {
                    "name": header_name,
                    "tais_code_date": header_date,
                    "filename": filename,
                    "sheetname": sheet_names,
                    "notes": notes,
                }
            )
        else:
            header_rec.name = header_name
            header_rec.filename = filename
            header_rec.sheetname = sheet_names
            header_rec.notes = notes
            header_rec.item_ids.unlink()

        date_str = header_date.strftime("%Y-%m-%d")
        for item in temp_data:
            item_model.create(
                {
                    "name": f"{item['tais_code']}:{date_str}",
                    "tais_code": item["tais_code"],
                    "manufacturer": item["manufacturer"],
                    "product_name": item["product_name"],
                    "model_number": item["model_number"],
                    "average_price": item["average_price"],
                    "price_cap": item["price_cap"],
                    "pricelist_id": header_rec.id,
                }
            )

        return {
            "type": "ir.actions.act_window",
            "res_model": "taisplus.pricelist",
            "view_mode": "form",
            "res_id": header_rec.id,
            "target": "current",
        }
